from __future__ import annotations

import html
import json
import os
import re
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from math import log10
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = next(ROOT.glob("*.xlsx"))
OUTPUT = ROOT / "assets" / "app-data.js"
CACHE = ROOT / "assets" / "bilibili-cache.json"

GOOD_TITLE_WORDS = [
    "\u671f\u672b",
    "\u590d\u4e60",
    "\u901f\u6210",
    "\u6559\u7a0b",
    "\u57fa\u7840",
    "\u5168\u5957",
    "\u5b8c\u6574",
    "\u7cbe\u8bb2",
    "\u516c\u5f00\u8bfe",
    "\u4e0d\u6302\u79d1",
    "\u8003\u524d",
]

BAD_TITLE_WORDS = [
    "\u7b54\u6848",
    "\u5e26\u5199",
    "\u4ee3\u5199",
    "\u6bd5\u8bbe",
    "\u8bfe\u8bbe",
    "\u5b89\u88c5",
    "\u6f14\u793a",
]

GENERIC_MARKERS = {"C\u8bed\u8a00", "Python", "Java"}

SUBJECT_RULES = [
    ("C\u8bed\u8a00", ["C\u8bed\u8a00 \u671f\u672b\u590d\u4e60", "C\u8bed\u8a00 \u7a0b\u5e8f\u8bbe\u8ba1 \u6559\u7a0b"]),
    ("Java Web", ["Java Web \u671f\u672b\u590d\u4e60", "Java Web \u5168\u5957\u6559\u7a0b"]),
    ("Java", ["Java \u671f\u672b\u590d\u4e60", "Java \u7a0b\u5e8f\u8bbe\u8ba1 \u6559\u7a0b"]),
    ("Python", ["Python \u7a0b\u5e8f\u8bbe\u8ba1 \u671f\u672b\u590d\u4e60", "Python \u96f6\u57fa\u7840 \u6559\u7a0b"]),
    ("\u6570\u636e\u7ed3\u6784", ["\u6570\u636e\u7ed3\u6784 \u671f\u672b\u590d\u4e60", "\u6570\u636e\u7ed3\u6784 \u5168\u5957\u6559\u7a0b"]),
    ("\u79bb\u6563\u6570\u5b66", ["\u79bb\u6563\u6570\u5b66 \u671f\u672b\u590d\u4e60", "\u79bb\u6563\u6570\u5b66 \u6559\u7a0b"]),
    ("\u9ad8\u7b49\u6570\u5b66", ["\u9ad8\u7b49\u6570\u5b66 \u671f\u672b\u590d\u4e60", "\u9ad8\u7b49\u6570\u5b66 \u6559\u7a0b"]),
    ("\u7ebf\u6027\u4ee3\u6570", ["\u7ebf\u6027\u4ee3\u6570 \u671f\u672b\u590d\u4e60", "\u7ebf\u6027\u4ee3\u6570 \u6559\u7a0b"]),
    ("\u6982\u7387\u8bba", ["\u6982\u7387\u8bba\u4e0e\u6570\u7406\u7edf\u8ba1 \u671f\u672b\u590d\u4e60", "\u6982\u7387\u8bba \u6559\u7a0b"]),
    ("\u5927\u5b66\u7269\u7406", ["\u5927\u5b66\u7269\u7406 \u671f\u672b\u590d\u4e60", "\u5927\u5b66\u7269\u7406 \u6559\u7a0b"]),
    ("\u64cd\u4f5c\u7cfb\u7edf", ["\u64cd\u4f5c\u7cfb\u7edf \u671f\u672b\u590d\u4e60", "\u64cd\u4f5c\u7cfb\u7edf \u6559\u7a0b"]),
    ("\u8ba1\u7b97\u673a\u7f51\u7edc", ["\u8ba1\u7b97\u673a\u7f51\u7edc \u671f\u672b\u590d\u4e60", "\u8ba1\u7b97\u673a\u7f51\u7edc \u6559\u7a0b"]),
    ("\u6570\u636e\u5e93", ["\u6570\u636e\u5e93 \u671f\u672b\u590d\u4e60", "\u6570\u636e\u5e93 SQL \u6559\u7a0b"]),
    ("MySQL", ["MySQL \u5168\u5957\u6559\u7a0b", "MySQL \u6570\u636e\u5e93 \u671f\u672b\u590d\u4e60"]),
    ("\u8f6f\u4ef6\u5de5\u7a0b", ["\u8f6f\u4ef6\u5de5\u7a0b \u671f\u672b\u590d\u4e60", "\u8f6f\u4ef6\u5de5\u7a0b \u6559\u7a0b"]),
    ("\u7f16\u8bd1\u539f\u7406", ["\u7f16\u8bd1\u539f\u7406 \u671f\u672b\u590d\u4e60", "\u7f16\u8bd1\u539f\u7406 \u6559\u7a0b"]),
    ("\u8ba1\u7b97\u673a\u7ec4\u6210", ["\u8ba1\u7b97\u673a\u7ec4\u6210\u539f\u7406 \u671f\u672b\u590d\u4e60", "\u8ba1\u7b97\u673a\u7ec4\u6210\u539f\u7406 \u6559\u7a0b"]),
    ("\u6570\u5b57\u7535\u5b50", ["\u6570\u5b57\u7535\u5b50\u6280\u672f \u671f\u672b\u590d\u4e60", "\u6570\u5b57\u7535\u5b50\u6280\u672f \u6559\u7a0b"]),
    ("\u6570\u5b57\u7535\u8def", ["\u6570\u5b57\u7535\u8def \u671f\u672b\u590d\u4e60", "\u6570\u5b57\u7535\u8def \u6559\u7a0b"]),
    ("\u6a21\u62df\u7535\u5b50", ["\u6a21\u62df\u7535\u5b50\u6280\u672f \u671f\u672b\u590d\u4e60", "\u6a21\u62df\u7535\u5b50\u6280\u672f \u6559\u7a0b"]),
    ("\u7535\u8def", ["\u7535\u8def\u5206\u6790 \u671f\u672b\u590d\u4e60", "\u7535\u8def \u6559\u7a0b"]),
    ("\u4fe1\u53f7\u4e0e\u7cfb\u7edf", ["\u4fe1\u53f7\u4e0e\u7cfb\u7edf \u671f\u672b\u590d\u4e60", "\u4fe1\u53f7\u4e0e\u7cfb\u7edf \u6559\u7a0b"]),
    ("\u901a\u4fe1\u539f\u7406", ["\u901a\u4fe1\u539f\u7406 \u671f\u672b\u590d\u4e60", "\u901a\u4fe1\u539f\u7406 \u6559\u7a0b"]),
    ("\u81ea\u52a8\u63a7\u5236", ["\u81ea\u52a8\u63a7\u5236\u539f\u7406 \u671f\u672b\u590d\u4e60", "\u81ea\u52a8\u63a7\u5236\u539f\u7406 \u6559\u7a0b"]),
    ("MATLAB", ["MATLAB \u96f6\u57fa\u7840 \u6559\u7a0b", "MATLAB \u671f\u672b\u590d\u4e60"]),
    ("\u5355\u7247\u673a", ["\u5355\u7247\u673a \u671f\u672b\u590d\u4e60", "\u5355\u7247\u673a \u6559\u7a0b"]),
    ("\u673a\u5668\u5b66\u4e60", ["\u673a\u5668\u5b66\u4e60 \u5168\u5957\u6559\u7a0b", "\u673a\u5668\u5b66\u4e60 \u671f\u672b\u590d\u4e60"]),
    ("\u6df1\u5ea6\u5b66\u4e60", ["\u6df1\u5ea6\u5b66\u4e60 \u5168\u5957\u6559\u7a0b", "\u6df1\u5ea6\u5b66\u4e60 \u671f\u672b\u590d\u4e60"]),
    ("OpenCV", ["OpenCV Python \u5168\u5957\u6559\u7a0b", "OpenCV \u8ba1\u7b97\u673a\u89c6\u89c9 \u6559\u7a0b"]),
    ("\u5927\u6570\u636e", ["\u5927\u6570\u636e \u671f\u672b\u590d\u4e60", "\u5927\u6570\u636e \u5168\u5957\u6559\u7a0b"]),
]


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def clean_major(value) -> str:
    text = clean(value)
    text = text.replace("\uff08", "(").replace("\uff09", ")")
    text = re.sub(r"\s+", "", text)
    return text.replace("(", "\uff08").replace(")", "\uff09")


def strip_highlight(value: str) -> str:
    text = re.sub(r"</?em[^>]*>", "", value or "")
    return clean(html.unescape(text))


def is_real_book(title: str) -> bool:
    if not title:
        return False
    skip_words = ["\u65e0\u9700\u6559\u6750", "\u7528\u4e0a\u5b66\u671f\u7684\u6559\u6750"]
    return not any(word in title for word in skip_words)


def simplify_keyword(value: str) -> str:
    text = clean(value)
    text = re.sub(r"[\uff08(][^\uff09)]*(?:\u7b2c|Edition|edition|版|慕课|微课|双色|英文|修订|AIGC).*?[\uff09)]", " ", text)
    text = re.sub(r"\u7b2c[一二三四五六七八九十0-9]+版|\([^)]+\)|（[^）]+）", " ", text)
    text = re.sub(r"[《》“”\"'、，,。:：；;|/\\\[\]【】\-_——]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def duration_seconds(value: str) -> int:
    parts = [part for part in clean(value).split(":") if part.isdigit()]
    if not parts:
        return 0
    total = 0
    for part in parts:
        total = total * 60 + int(part)
    return total


def search_terms_for(row: dict) -> list[str]:
    base = f"{row['title']} {row['course']}".strip()
    terms = [base]

    simple_title = simplify_keyword(row["title"])
    simple_course = simplify_keyword(row["course"])
    if len(simple_title) >= 2:
        terms.append(simple_title)

    combined = f"{row['title']} {row['course']}"
    for marker, marker_terms in SUBJECT_RULES:
        if marker.lower() in combined.lower():
            terms.extend(marker_terms)

    for term in [simple_course, f"{simple_title} {simple_course}".strip()]:
        if len(term) >= 2:
            terms.append(term)

    unique_terms = []
    for term in terms:
        term = clean(term)
        if term and term not in unique_terms:
            unique_terms.append(term)
    return unique_terms[:5]


def video_score(video: dict, row: dict, query: str) -> float:
    title = video["title"]
    title_lower = title.lower()
    play = int(video.get("play") or 0)
    seconds = duration_seconds(video.get("duration", ""))

    score = log10(max(play, 1)) * 18
    score += min(seconds / 3600, 10) * 12

    for word in GOOD_TITLE_WORDS:
        if word in title:
            score += 18
    for word in BAD_TITLE_WORDS:
        if word in title:
            score -= 45

    source = f"{row['title']} {row['course']}"
    matched_markers = [
        marker for marker, _terms in SUBJECT_RULES
        if marker.lower() in source.lower()
    ]
    specific_markers = [marker for marker in matched_markers if marker not in GENERIC_MARKERS]
    for marker in specific_markers:
        if marker.lower() in title_lower:
            score += 55
        else:
            score -= 70
    if not specific_markers:
        for marker in matched_markers:
            if marker.lower() in title_lower:
                score += 35

    course_key = simplify_keyword(row["course"])
    if course_key and course_key.lower() in title_lower:
        score += 20

    if seconds < 600:
        score -= 35
    elif seconds > 1800:
        score += 15

    if query == f"{row['title']} {row['course']}".strip():
        score += 8
    return score


def normalize_video(item: dict) -> dict | None:
    if item.get("type") != "video" or not item.get("bvid"):
        return None
    title = strip_highlight(item.get("title", ""))
    if not title:
        return None
    return {
        "title": title,
        "author": clean(item.get("author", "")),
        "bvid": clean(item.get("bvid", "")),
        "url": f"https://www.bilibili.com/video/{item.get('bvid')}",
        "play": int(item.get("play") or 0),
        "duration": clean(item.get("duration", "")),
    }


def fetch_bilibili_videos(query: str) -> list[dict]:
    keyword = urllib.parse.quote(query)
    url = (
        "https://api.bilibili.com/x/web-interface/search/type"
        f"?search_type=video&keyword={keyword}&page=1"
    )
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
            ),
            "Referer": "https://search.bilibili.com/",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8", errors="ignore"))
    except Exception:
        return []

    videos = []
    for item in payload.get("data", {}).get("result", [])[:20]:
        video = normalize_video(item)
        if video:
            videos.append(video)
    return videos


def search_bilibili_for_row(row: dict) -> dict | None:
    candidates = []
    for query in search_terms_for(row):
        for video in fetch_bilibili_videos(query):
            video = {**video, "matched_query": query}
            candidates.append((video_score(video, row, query), video))

    if not candidates:
        return None

    candidates.sort(key=lambda item: item[0], reverse=True)
    best = candidates[0][1]
    return {
        "title": best["title"],
        "author": best["author"],
        "bvid": best["bvid"],
        "url": best["url"],
        "play": best["play"],
        "duration": best["duration"],
        "matched_query": best["matched_query"],
    }


def load_rows() -> list[dict]:
    workbook = load_workbook(WORKBOOK, data_only=True)
    rows = []
    seen_rows = set()

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            major = clean_major(row[0])
            year = clean(row[1])
            course = clean(row[2])
            title = clean(row[3])
            author = clean(row[4] if len(row) > 4 else "")
            publisher = clean(row[5] if len(row) > 5 else "")
            if not (major and year and course and is_real_book(title)):
                continue
            key = (major, year, course, title, author, publisher)
            if key in seen_rows:
                continue
            seen_rows.add(key)
            rows.append(
                {
                    "majorName": major,
                    "year": year,
                    "course": course,
                    "title": title,
                    "author": author,
                    "publisher": publisher,
                }
            )
    return rows


def load_cache() -> dict:
    if not CACHE.exists():
        return {}
    try:
        return json.loads(CACHE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def write_cache(cache: dict) -> None:
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    rows = load_rows()

    major_order = []
    for row in rows:
        if row["majorName"] not in major_order:
            major_order.append(row["majorName"])

    majors = []
    major_ids = {}
    for index, name in enumerate(major_order, start=1):
        major_rows = [row for row in rows if row["majorName"] == name]
        year_count = len({row["year"] for row in major_rows})
        book_count = len(major_rows)
        major_id = f"major-{index}"
        major_ids[name] = major_id
        majors.append(
            {
                "id": major_id,
                "name": name,
                "desc": f"{year_count} \u4e2a\u5e74\u7ea7 \u00b7 {book_count} \u672c\u6559\u6750",
            }
        )

    cache = load_cache()
    unique_rows = {}
    for row in rows:
        query = f"{row['title']} {row['course']}".strip()
        if query not in unique_rows:
            unique_rows[query] = row
        if query not in cache:
            cache[query] = None

    refresh = os.environ.get("REFRESH_BILI_SEARCH")
    pending = [] if os.environ.get("SKIP_BILI_SEARCH") else [
        query for query in unique_rows if refresh or query not in cache or cache.get(query) is None
    ]
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            executor.submit(search_bilibili_for_row, unique_rows[query]): query
            for query in pending
        }
        for index, future in enumerate(as_completed(futures), start=1):
            query = futures[future]
            try:
                cache[query] = future.result()
            except Exception:
                cache[query] = None
            if index % 40 == 0:
                write_cache(cache)
    write_cache(cache)

    books = []
    for index, row in enumerate(rows, start=1):
        query = f"{row['title']} {row['course']}".strip()
        video = cache.get(query)
        routes = []
        if video:
            routes.append(
                {
                    "title": "B\u7ad9\u76f8\u5173\u89c6\u9891",
                    "search": (
                        f"\u89c6\u9891\uff1a{video['title']}\uff5cUP\uff1a{video['author']}\uff5c"
                        f"{video['bvid']}\uff5c\u64ad\u653e\uff1a{video['play']}"
                        f"\uff5c\u641c\u7d22\uff1a{video.get('matched_query', query)}"
                    ),
                    "url": video["url"],
                    "query": query,
                }
            )
        routes.append(
            {
                "title": "B\u7ad9\u641c\u7d22\u5173\u952e\u8bcd",
                "search": f"\u5efa\u8bae\u641c\u7d22\uff1a{query}",
                "query": query,
            }
        )

        tags = [value for value in [row["author"], row["publisher"]] if value][:2]
        note_parts = []
        if row["author"]:
            note_parts.append(f"\u4f5c\u8005\uff1a{row['author']}")
        if row["publisher"]:
            note_parts.append(f"\u51fa\u7248\u793e\uff1a{row['publisher']}")
        note_parts.append("\u6309 Excel \u6559\u6750\u8868\u6c47\u603b\u5c55\u793a\u3002")

        books.append(
            {
                "id": f"book-{index}",
                "major": major_ids[row["majorName"]],
                "year": row["year"],
                "course": row["course"],
                "title": row["title"],
                "tags": tags,
                "note": "\uff1b".join(note_parts),
                "routes": routes,
            }
        )

    payload = {"majors": majors, "books": books}
    OUTPUT.write_text(
        "window.APP_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    found = sum(1 for video in cache.values() if video)
    print(f"majors={len(majors)} books={len(books)} unique_queries={len(unique_rows)} found={found}")


if __name__ == "__main__":
    main()
